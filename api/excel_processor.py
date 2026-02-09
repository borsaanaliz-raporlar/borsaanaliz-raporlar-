# /api/excel_processor.py - GÜNCEL EXCEL BUL ve 3 SAYFA OKU
import urllib.request
import tempfile
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
import json
import os
from typing import Dict, List, Any
import hashlib
import pickle

class ExcelProcessor:
    """GÜNCEL Excel bul ve 3 sayfa oku"""
    
    def __init__(self):
        self.cache_dir = "/tmp/borsa_cache"
        self.cache_duration = 7200  # 2 saat cache
        os.makedirs(self.cache_dir, exist_ok=True)
    
    def find_latest_excel(self) -> tuple:
        """EN GÜNCEL Excel dosyasını bul"""
        try:
            base_url = "https://borsaanaliz-raporlar.vercel.app/raporlar/"
            today = datetime.now()
            
            # Son 7 günü kontrol et
            for i in range(7):
                date = today - timedelta(days=i)
                date_str = date.strftime("%d%m%Y")
                filename = f"BORSAANALIZ_V11_TAM_{date_str}.xlsm"
                file_url = f"{base_url}{filename}"
                
                try:
                    # HEAD isteği ile dosya var mı kontrol et
                    req = urllib.request.Request(file_url, method='HEAD')
                    with urllib.request.urlopen(req, timeout=10) as response:
                        if response.status == 200:
                            print(f"✅ GÜNCEL EXCEL BULUNDU: {filename}")
                            
                            # Tarihi çıkar
                            date_match = re.search(r'(\d{2})(\d{2})(\d{4})\.xlsm$', filename)
                            if date_match:
                                day, month, year = date_match.groups()
                                file_date = datetime(int(year), int(month), int(day))
                                return file_url, file_date.strftime("%d.%m.%Y")
                            else:
                                return file_url, "güncel"
                except:
                    continue  # Bu dosya yok, diğerini dene
            
            # Hiçbiri yoksa fallback
            print("⚠️ Güncel dosya bulunamadı, fallback kullanılıyor...")
            return "https://borsaanaliz-raporlar.vercel.app/raporlar/BORSAANALIZ_V11_TAM_06022026.xlsm", "06.02.2026"
            
        except Exception as e:
            print(f"❌ Excel bulma hatası: {e}")
            return "https://borsaanaliz-raporlar.vercel.app/raporlar/BORSAANALIZ_V11_TAM_06022026.xlsm", "06.02.2026"
    
    def get_cache_key(self, excel_url: str) -> str:
        """Cache key oluştur"""
        url_hash = hashlib.md5(excel_url.encode()).hexdigest()
        return f"{self.cache_dir}/excel_data_{url_hash}.pkl"
    
    def is_cache_valid(self, cache_file: str) -> bool:
        """Cache geçerli mi?"""
        if not os.path.exists(cache_file):
            return False
        
        file_age = datetime.now().timestamp() - os.path.getmtime(cache_file)
        return file_age < self.cache_duration
    
    def load_from_cache(self, cache_file: str) -> Dict:
        """Cache'den yükle"""
        try:
            with open(cache_file, 'rb') as f:
                return pickle.load(f)
        except:
            return None
    
    def save_to_cache(self, cache_file: str, data: Dict):
        """Cache'e kaydet"""
        try:
            with open(cache_file, 'wb') as f:
                pickle.dump(data, f)
        except:
            pass
    
    def download_excel(self, excel_url: str) -> bytes:
        """Excel'i indir"""
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        req = urllib.request.Request(excel_url, headers=headers)
        
        with urllib.request.urlopen(req, timeout=60) as response:  # 60 saniye
            return response.read()
    
    def clean_header(self, header):
        """Başlığı temizle: 'Hisse (06-02-2026)' -> 'Hisse'"""
        if not header:
            return ""
        header = str(header).split('(')[0].strip()
        header = re.sub(r'\s+', ' ', header)
        return header
    
    def parse_cell_value(self, value):
        """Hücre değerini parse et"""
        if value is None:
            return None
        
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y")
        elif isinstance(value, float):
            # Ondalık hassasiyet
            return round(value, 4) if abs(value) < 1000 else round(value, 2)
        elif isinstance(value, (int, float)):
            return float(value) if '.' in str(value) else int(value)
        else:
            return str(value).strip()
    
    def read_excel_data(self, excel_url: str = None) -> Dict:
        """3 SAYFA TAM OKU - GÜNCEL EXCEL"""
        # Eğer URL verilmediyse, en günceli bul
        if not excel_url:
            print("🔍 En güncel Excel aranıyor...")
            excel_url, excel_date = self.find_latest_excel()
            print(f"✅ Bulunan: {os.path.basename(excel_url)} ({excel_date})")
        else:
            excel_date = "manuel_girildi"
        
        print(f"📊 3 SAYFA Excel işleniyor: {excel_url}")
        
        # 1. Cache kontrolü
        cache_file = self.get_cache_key(excel_url)
        if self.is_cache_valid(cache_file):
            print("✅ Cache'den yükleniyor...")
            cached_data = self.load_from_cache(cache_file)
            if cached_data:
                return cached_data
        
        start_time = datetime.now()
        
        try:
            # 2. Excel'i indir
            print("⬇️  Excel indiriliyor...")
            excel_content = self.download_excel(excel_url)
            
            # 3. Geçici dosyaya yaz
            with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            # 4. openpyxl ile aç (read_only modunda - HIZLI)
            print("📖 Excel açılıyor...")
            wb = load_workbook(tmp_path, data_only=True, read_only=True)
            
            print(f"✅ Excel açıldı. Sayfalar: {wb.sheetnames}")
            
            result = {
                "excel_url": excel_url,
                "excel_date": excel_date,
                "timestamp": datetime.now().isoformat(),
                "sheets": {},
                "total_symbols": 0,
                "load_time": None
            }
            
            # 5. 1. SAYFA: SİNYALLER (630+ hisse)
            if "Sinyaller" in wb.sheetnames:
                print("📈 Sinyaller sayfası TAM okunuyor...")
                ws = wb["Sinyaller"]
                
                # Başlıkları oku
                headers = []
                for col in range(1, 150):  # 150 sütun
                    cell_val = ws.cell(row=1, column=col).value
                    if not cell_val:
                        break
                    headers.append(self.clean_header(cell_val))
                
                print(f"📋 Sinyaller: {len(headers)} sütun bulundu")
                
                # TÜM hisseleri oku (ilk 1000 satır = 630+ hisse)
                sinyaller_data = {}
                row_count = 0
                
                for row in ws.iter_rows(min_row=2, max_row=1001, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    hisse_adi = str(row[0]).strip()
                    if not hisse_adi:
                        continue
                    
                    hisse_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            cell_val = row[col_idx]
                            if cell_val is not None:
                                hisse_dict[header] = self.parse_cell_value(cell_val)
                    
                    sinyaller_data[hisse_adi] = hisse_dict
                    row_count += 1
                    
                    if row_count % 200 == 0:
                        print(f"   ...{row_count} hisse okundu")
                
                result["sheets"]["Sinyaller"] = {
                    "headers": headers,
                    "hisseler": sinyaller_data,
                    "toplam_hisse": len(sinyaller_data),
                    "okunan_satir": row_count
                }
                result["total_symbols"] += len(sinyaller_data)
                print(f"✅ Sinyaller: {len(sinyaller_data)} hisse okundu")
            
            # 6. 2. SAYFA: ENDEKSLER
            if "ENDEKSLER" in wb.sheetnames:
                print("📊 ENDEKSLER sayfası TAM okunuyor...")
                ws = wb["ENDEKSLER"]
                
                # Başlıkları oku
                headers = []
                for col in range(1, 100):
                    cell_val = ws.cell(row=1, column=col).value
                    if not cell_val:
                        break
                    headers.append(self.clean_header(cell_val))
                
                # TÜM endeksleri oku (ilk 200 satır)
                endeks_data = {}
                row_count = 0
                
                for row in ws.iter_rows(min_row=2, max_row=201, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    sembol_adi = str(row[0]).strip()
                    if not sembol_adi:
                        continue
                    
                    sembol_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            cell_val = row[col_idx]
                            if cell_val is not None:
                                sembol_dict[header] = self.parse_cell_value(cell_val)
                    
                    endeks_data[sembol_adi] = sembol_dict
                    row_count += 1
                
                result["sheets"]["ENDEKSLER"] = {
                    "headers": headers,
                    "semboller": endeks_data,
                    "toplam_sembol": len(endeks_data),
                    "okunan_satir": row_count
                }
                result["total_symbols"] += len(endeks_data)
                print(f"✅ ENDEKSLER: {len(endeks_data)} sembol okundu")
            
            # 7. 3. SAYFA: FON_EMTIA_COIN_DOVIZ
            if "FON_EMTIA_COIN_DOVIZ" in wb.sheetnames:
                print("💰 FON_EMTIA_COIN_DOVIZ sayfası TAM okunuyor...")
                ws = wb["FON_EMTIA_COIN_DOVIZ"]
                
                # Başlıkları oku
                headers = []
                for col in range(1, 100):
                    cell_val = ws.cell(row=1, column=col).value
                    if not cell_val:
                        break
                    headers.append(self.clean_header(cell_val))
                
                # TÜM sembolleri oku (ilk 150 satır)
                fon_data = {}
                row_count = 0
                
                for row in ws.iter_rows(min_row=2, max_row=151, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    sembol_adi = str(row[0]).strip()
                    if not sembol_adi:
                        continue
                    
                    sembol_dict = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            cell_val = row[col_idx]
                            if cell_val is not None:
                                sembol_dict[header] = self.parse_cell_value(cell_val)
                    
                    fon_data[sembol_adi] = sembol_dict
                    row_count += 1
                
                result["sheets"]["FON_EMTIA_COIN_DOVIZ"] = {
                    "headers": headers,
                    "semboller": fon_data,
                    "toplam_sembol": len(fon_data),
                    "okunan_satir": row_count
                }
                result["total_symbols"] += len(fon_data)
                print(f"✅ FON_EMTIA_COIN_DOVIZ: {len(fon_data)} sembol okundu")
            
            # 8. Temizlik
            wb.close()
            os.unlink(tmp_path)
            
            # 9. Metadata
            load_time = (datetime.now() - start_time).total_seconds()
            result["load_time"] = load_time
            
            print(f"🎉 3 SAYFA TAM OKUNDU! Toplam: {result['total_symbols']} sembol, {load_time:.2f}s")
            
            # 10. Cache'e kaydet
            self.save_to_cache(cache_file, result)
            
            return result
            
        except Exception as e:
            print(f"❌ Excel okuma hatası: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def get_available_fields(self, excel_data: Dict, sheet_name: str = "Sinyaller") -> List[str]:
        """Mevcut teknik alanları listele"""
        if sheet_name not in excel_data.get("sheets", {}):
            return []
        
        sheet_data = excel_data["sheets"][sheet_name]
        if "hisseler" in sheet_data:
            first_hisse = next(iter(sheet_data["hisseler"].values()), {})
        elif "semboller" in sheet_data:
            first_hisse = next(iter(sheet_data["semboller"].values()), {})
        else:
            return []
        
        return list(first_hisse.keys())
    
    def get_all_symbols(self, excel_data: Dict) -> Dict:
        """Tüm sembolleri listele (3 sayfadan)"""
        symbols = {
            "Sinyaller": [],
            "ENDEKSLER": [],
            "FON_EMTIA_COIN_DOVIZ": []
        }
        
        for sheet_name in symbols.keys():
            if sheet_name in excel_data.get("sheets", {}):
                sheet_data = excel_data["sheets"][sheet_name]
                if "hisseler" in sheet_data:
                    symbols[sheet_name] = list(sheet_data["hisseler"].keys())[:20]  # İlk 20
                elif "semboller" in sheet_data:
                    symbols[sheet_name] = list(sheet_data["semboller"].keys())[:20]
        
        return symbols

# Global instance
excel_processor = ExcelProcessor()
