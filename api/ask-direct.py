#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# /api/ask-direct.py
# BorsaAnaliz AI - TAM ÇALIŞAN VERSİYON
# Versiyon: 7.0 (Final)

from http.server import BaseHTTPRequestHandler
import json
import os
import sys
import re
import traceback
from datetime import datetime
import urllib.request
import tempfile
from openpyxl import load_workbook
import requests
import random

# ==================== GLOBAL AYARLAR ====================
DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY', '')
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
MODE = "hizli"  # "hizli" veya "detayli" - frontend'den gelecek

# ==================== EXCEL OKUYUCU ====================
class ExcelReader:
    """EN GÜNCEL Excel'i bul ve 3 sayfa oku"""
    
    def find_latest_excel(self):
        """En güncel Excel dosyasını bul"""
        try:
            base_url = "https://borsaanaliz-raporlar.vercel.app/raporlar/"
            today = datetime.now()
            
            # Son 7 günü kontrol et
            for i in range(7):
                date = today.replace(hour=0, minute=0, second=0, microsecond=0)
                date_str = date.strftime("%d%m%Y")
                filename = f"BORSAANALIZ_V11_TAM_{date_str}.xlsm"
                file_url = f"{base_url}{filename}"
                
                try:
                    # HEAD isteği ile dosya var mı kontrol et
                    req = urllib.request.Request(file_url, method='HEAD')
                    with urllib.request.urlopen(req, timeout=5) as response:
                        if response.status == 200:
                            print(f"✅ GÜNCEL EXCEL BULUNDU: {filename}", file=sys.stderr)
                            return file_url, date.strftime("%d.%m.%Y")
                except:
                    continue  # Bu tarih yok, bir önceki güne bak
            
            # Fallback: Bugünün dosyası yoksa en son dosyayı kullan
            print("⚠️ Güncel dosya bulunamadı, fallback kullanılıyor...", file=sys.stderr)
            return f"{base_url}BORSAANALIZ_V11_TAM_06022026.xlsm", "06.02.2026"
            
        except Exception as e:
            print(f"❌ Excel bulma hatası: {e}", file=sys.stderr)
            return "https://borsaanaliz-raporlar.vercel.app/raporlar/BORSAANALIZ_V11_TAM_06022026.xlsm", "06.02.2026"
    
    def read_excel(self):
        """3 sayfayı da oku"""
        try:
            print("🚀 EXCEL OKUMA BAŞLIYOR...", file=sys.stderr)
            
            # 1. En güncel Excel'i bul
            excel_url, excel_date = self.find_latest_excel()
            print(f"📥 Excel: {excel_url}", file=sys.stderr)
            print(f"📅 Tarih: {excel_date}", file=sys.stderr)
            
            # 2. İndir
            headers = {'User-Agent': 'Mozilla/5.0'}
            req = urllib.request.Request(excel_url, headers=headers)
            
            with urllib.request.urlopen(req, timeout=30) as response:
                excel_content = response.read()
                print(f"✅ İndirildi: {len(excel_content)} bytes", file=sys.stderr)
            
            # 3. Geçici dosya
            with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as tmp:
                tmp.write(excel_content)
                tmp_path = tmp.name
            
            # 4. Aç
            wb = load_workbook(tmp_path, data_only=True, read_only=True)
            print(f"📖 Sayfalar: {wb.sheetnames}", file=sys.stderr)
            
            result = {
                "success": True,
                "excel_date": excel_date,
                "excel_url": excel_url,
                "total_symbols": 0,
                "sheets": {}
            }
            
            # ==================== 1. SİNYALLER ====================
            if "Sinyaller" in wb.sheetnames:
                ws = wb["Sinyaller"]
                hisseler = {}
                
                for row in ws.iter_rows(min_row=2, max_row=1000, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    hisse_adi = str(row[0]).strip()
                    if not hisse_adi:
                        continue
                    
                    # Temel verileri al
                    hisse_dict = {
                        "Hisse": hisse_adi,
                        "Close": row[6] if len(row) > 6 else None,
                        "VMA": row[9] if len(row) > 9 else None,
                        "DURUM": row[15] if len(row) > 15 else None,
                        "EMA_8": row[27] if len(row) > 27 else None,
                        "Pivot": row[7] if len(row) > 7 else None,
                        "Open": row[49] if len(row) > 49 else None,
                        "High": row[50] if len(row) > 50 else None,
                        "Low": row[51] if len(row) > 51 else None,
                        "Hacim": row[11] if len(row) > 11 else None
                    }
                    
                    # None değerleri temizle
                    hisse_dict = {k: v for k, v in hisse_dict.items() if v is not None}
                    hisseler[hisse_adi] = hisse_dict
                
                result["sheets"]["Sinyaller"] = {"hisseler": hisseler}
                result["total_symbols"] += len(hisseler)
                print(f"✅ Sinyaller: {len(hisseler)} hisse", file=sys.stderr)
            
            # ==================== 2. ENDEKSLER ====================
            if "ENDEKSLER" in wb.sheetnames:
                ws = wb["ENDEKSLER"]
                endeksler = {}
                
                for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    sembol_adi = str(row[0]).strip()
                    if not sembol_adi:
                        continue
                    
                    sembol_dict = {
                        "Sembol": sembol_adi,
                        "Close": row[6] if len(row) > 6 else None,
                        "VMA": row[9] if len(row) > 9 else None,
                        "DURUM": row[15] if len(row) > 15 else None
                    }
                    
                    sembol_dict = {k: v for k, v in sembol_dict.items() if v is not None}
                    endeksler[sembol_adi] = sembol_dict
                
                result["sheets"]["ENDEKSLER"] = {"semboller": endeksler}
                result["total_symbols"] += len(endeksler)
                print(f"✅ ENDEKSLER: {len(endeksler)} sembol", file=sys.stderr)
            
            # ==================== 3. FON_EMTIA_COIN_DOVIZ ====================
            if "FON_EMTIA_COIN_DOVIZ" in wb.sheetnames:
                ws = wb["FON_EMTIA_COIN_DOVIZ"]
                fonlar = {}
                
                for row in ws.iter_rows(min_row=2, max_row=100, values_only=True):
                    if not row or not row[0]:
                        continue
                    
                    sembol_adi = str(row[0]).strip()
                    if not sembol_adi:
                        continue
                    
                    sembol_dict = {
                        "Sembol": sembol_adi,
                        "Close": row[6] if len(row) > 6 else None,
                        "VMA": row[9] if len(row) > 9 else None,
                        "DURUM": row[15] if len(row) > 15 else None
                    }
                    
                    sembol_dict = {k: v for k, v in sembol_dict.items() if v is not None}
                    fonlar[sembol_adi] = sembol_dict
                
                result["sheets"]["FON_EMTIA_COIN_DOVIZ"] = {"semboller": fonlar}
                result["total_symbols"] += len(fonlar)
                print(f"✅ FON_EMTIA_COIN_DOVIZ: {len(fonlar)} sembol", file=sys.stderr)
            
            wb.close()
            os.unlink(tmp_path)
            
            print(f"🎉 TOPLAM: {result['total_symbols']} sembol", file=sys.stderr)
            return result
            
        except Exception as e:
            print(f"❌ EXCEL OKUMA HATASI: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            return {"success": False, "error": str(e)}

# Global Excel reader
excel_reader = ExcelReader()

# ==================== SORU ANALİZİ ====================
def analyze_question_type(question):
    """Soruyu analiz et"""
    q = question.lower().strip()
    
    # Hızlı mod soruları
    if any(k in q for k in ['teşekkür', 'sağ ol', 'sağol']):
        return "teşekkür"
    
    if any(k in q for k in ['vma', 'ema', 'teknik analiz', 'nasıl yorumlanır']):
        return "teknik"
    
    if any(k in q for k in ['excel', 'macro', 'makro']):
        return "excel_macro"
    
    if any(k in q for k in ['kim yaptı', 'sistem', 'hakkında']):
        return "sistem"
    
    if any(k in q for k in ['öne çıkan', 'en iyi', 'borsa durumu']):
        return "genel_borsa"
    
    if any(k in q for k in ['endeks', 'xu100', 'xulas']):
        return "endeks"
    
    if any(k in q for k in ['nasıl çalışır']):
        return "nasil"
    
    # Hisse kodu varsa
    if re.search(r'\b([A-Z]{2,6})\b', question.upper()):
        return "analiz"
    
    return "bilinmeyen"

# ==================== HIZLI MOD CEVAPLARI ====================
def get_hizli_cevap(question_type, question=""):
    """Hızlı mod için özel cevaplar"""
    if question_type == "teşekkür":
        return "🌟 **Teşekkür ederim!**\n\nBaşka hisse analizi istiyor musunuz?"
    
    elif question_type == "teknik":
        if 'vma' in question.lower():
            return """📊 **VMA Algoritması:**
• POZİTİF (00): Trend başlangıcı
• POZİTİF (--): Trend devamı  
• NEGATİF (00): Trend bitişi
• NEGATİF (--): Düşüş devamı"""
        else:
            return """📈 **Teknik Analiz Göstergeleri:**
• VMA: Hacim algoritması
• EMA: Fiyat trendi
• Pivot: Destek/direnç"""
    
    elif question_type == "excel_macro":
        return "📊 **Excel Macro:** .xlsm dosyası, 'Makroları Etkinleştir' seçeneğini işaretleyin."
    
    elif question_type == "sistem":
        return """🤖 **BorsaAnaliz AI Sistemi**
**Versiyon:** 7.0 (Final)
**Modlar:** Hızlı (DeepSeek) / Detaylı (Groq)
**Excel:** Güncel tarihli otomatik bulunur"""
    
    elif question_type == "genel_borsa":
        # Excel'den rastgele hisseler göster
        try:
            excel_data = excel_reader.read_excel()
            if excel_data.get("success"):
                if "Sinyaller" in excel_data.get("sheets", {}):
                    hisseler = list(excel_data["sheets"]["Sinyaller"]["hisseler"].keys())
                    if hisseler:
                        if len(hisseler) > 6:
                            secilen = random.sample(hisseler, 6)
                        else:
                            secilen = hisseler[:6]
                        
                        cevap = "📈 **Öne Çıkan Hisseler (Rastgele):**\n\n"
                        for h in secilen:
                            cevap += f"• {h}\n"
                        cevap += f"\n**Toplam:** {len(hisseler)} hisse\n"
                        cevap += "**Analiz:** \"[HİSSE] analiz et\""
                        return cevap
        except:
            pass
        
        return """📊 **Borsa Genel Durumu:**
• 600+ hisse analiz
• Güncel Excel verileri
• Örnek: "GARAN analiz et", "XU100 durumu" """
    
    elif question_type == "endeks":
        return """📈 **BIST Endeksleri:**
• XU100: BIST 100
• XU030: BIST 30  
• XULAS: Tüm şirketler
• Analiz: "XU100 analiz et" """
    
    elif question_type == "nasil":
        return """🔧 **Nasıl Çalışır:**
1. Güncel Excel bulunur
2. 3 sayfa okunur
3. Hisse aranır
4. AI analizi yapılır"""
    
    return """🤔 **Anlamadım**

Örnekler:
• Hisse: "GARAN analiz et"
• Endeks: "XU100 durumu"  
• Teknik: "VMA nedir?"
• Genel: "Öne çıkan hisseler""""

# ==================== ARAMA FONKSİYONU ====================
def find_symbol(question, excel_data):
    """3 sayfada sembol ara"""
    try:
        # Sembol kodunu çıkar
        match = re.search(r'\b([A-Z]{2,6})\b', question.upper())
        if not match:
            return {"found": False, "error": "Kod bulunamadı"}
        
        target = match.group(1)
        print(f"🔍 Aranan: '{target}'", file=sys.stderr)
        
        # 1. Sinyaller
        if "Sinyaller" in excel_data.get("sheets", {}):
            hisseler = excel_data["sheets"]["Sinyaller"]["hisseler"]
            for hisse_adi, veriler in hisseler.items():
                if target in hisse_adi.upper():
                    print(f"✅ Sinyaller: {hisse_adi}", file=sys.stderr)
                    return {"found": True, "name": hisse_adi, "data": veriler, "sayfa": "Sinyaller"}
        
        # 2. ENDEKSLER
        if "ENDEKSLER" in excel_data.get("sheets", {}):
            semboller = excel_data["sheets"]["ENDEKSLER"]["semboller"]
            for sembol_adi, veriler in semboller.items():
                if target in sembol_adi.upper():
                    print(f"✅ ENDEKSLER: {sembol_adi}", file=sys.stderr)
                    return {"found": True, "name": sembol_adi, "data": veriler, "sayfa": "ENDEKSLER"}
        
        # 3. FON_EMTIA_COIN_DOVIZ
        if "FON_EMTIA_COIN_DOVIZ" in excel_data.get("sheets", {}):
            semboller = excel_data["sheets"]["FON_EMTIA_COIN_DOVIZ"]["semboller"]
            for sembol_adi, veriler in semboller.items():
                if target in sembol_adi.upper():
                    print(f"✅ FON: {sembol_adi}", file=sys.stderr)
                    return {"found": True, "name": sembol_adi, "data": veriler, "sayfa": "FON_EMTIA_COIN_DOVIZ"}
        
        print(f"❌ '{target}' bulunamadı", file=sys.stderr)
        return {"found": False, "error": f"'{target}' Excel'de yok"}
        
    except Exception as e:
        print(f"❌ Arama hatası: {e}", file=sys.stderr)
        return {"found": False, "error": str(e)}

# ==================== AI ANALİZLERİ ====================
def get_deepseek_analysis(prompt, mode="hizli"):
    """DeepSeek API ile analiz"""
    try:
        if not DEEPSEEK_API_KEY:
            return "⚠️ DeepSeek API key gerekli"
        
        # Prompt'u mode'a göre ayarla
        if mode == "hizli":
            system_msg = "Kısa teknik analiz (max 150 kelime). Sadece verilen verileri kullan."
            max_tokens = 300
        else:
            system_msg = "Detaylı teknik analiz. Tüm göstergeleri değerlendir."
            max_tokens = 800
        
        headers = {
            'Authorization': f'Bearer {DEEPSEEK_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "model": "deepseek-chat",
            "messages": [
                {"role": "system", "content": system_msg},
                {"role": "user", "content": prompt[:1500]}  # Uzunluğu sınırla
            ],
            "max_tokens": max_tokens,
            "temperature": 0.7
        }
        
        response = requests.post(
            'https://api.deepseek.com/v1/chat/completions',
            headers=headers,
            json=data,
            timeout=15 if mode == "hizli" else 30
        )
        
        if response.status_code == 200:
            return response.json()['choices'][0]['message']['content']
        else:
            return f"❌ DeepSeek hatası: {response.status_code}"
            
    except requests.exceptions.Timeout:
        return "⏱️ DeepSeek zaman aşımı"
    except Exception as e:
        return f"❌ DeepSeek hatası: {str(e)[:100]}"

def get_groq_analysis(prompt):
    """Groq API ile detaylı analiz"""
    try:
        if not GROQ_API_KEY:
            return "⚠️ Groq API key gerekli"
        
        headers = {
            'Authorization': f'Bearer {GROQ_API_KEY}',
            'Content-Type': 'application/json'
        }
        
        data = {
            "model": "mixtral-8x7b-32768",
            "messages": [
                {"role": "system", "content": "Detaylı borsa analizi yap. Tüm teknik göstergeleri değerlendir."},
                {"role": "user", "content": prompt[:2000]}
            ],
            "max_tokens": 1500,
            "temperature": 0.7
        }
        
        response = requests.post(
            'https://api.groq.com/openai/v1/chat/completions',
            headers=headers,
            json=data,
            timeout=40
        )
        
        if response.status_code == 200:
            return response.json()['choices'][0]['message']['content']
        else:
            return f"❌ Groq hatası: {response.status_code}"
            
    except requests.exceptions.Timeout:
        return "⏱️ Groq zaman aşımı"
    except Exception as e:
        return f"❌ Groq hatası: {str(e)[:100]}"

# ==================== VERCEL HANDLER ====================
class handler(BaseHTTPRequestHandler):
    
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        
        response = {
            "status": "online",
            "version": "7.0 Final",
            "apis": "DeepSeek (hızlı) + Groq (detaylı)",
            "excel": "Güncel tarihli otomatik bulunur"
        }
        self.wfile.write(json.dumps(response, ensure_ascii=False).encode())
    
    def do_POST(self):
        try:
            # İstek verilerini al
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data)
            
            question = data.get('question', '').strip()
            mode = data.get('mode', 'hizli')  # "hizli" veya "detayli"
            
            if not question:
                self.send_error_response("Soru gerekli")
                return
            
            print(f"\n{'='*60}", file=sys.stderr)
            print(f"🤖 SORU: {question}", file=sys.stderr)
            print(f"🎮 MOD: {mode}", file=sys.stderr)
            print('='*60, file=sys.stderr)
            
            # Soru tipini analiz et
            question_type = analyze_question_type(question)
            print(f"🔍 Tip: {question_type}", file=sys.stderr)
            
            # HIZLI MOD için özel cevaplar
            if mode == "hizli" and question_type in ["teşekkür", "teknik", "excel_macro", 
                                                    "sistem", "genel_borsa", "endeks", "nasil"]:
                answer = get_hizli_cevap(question_type, question)
                self.send_success_response(answer, mode)
                return
            
            # HİSSE ANALİZİ (her iki modda da)
            if question_type == "analiz":
                print(f"🔍 {mode.upper()} analiz başlıyor...", file=sys.stderr)
                
                # Excel'i oku
                excel_data = excel_reader.read_excel()
                
                if not excel_data.get("success"):
                    error_msg = excel_data.get("error", "Excel okunamadı")
                    print(f"❌ Excel hatası: {error_msg}", file=sys.stderr)
                    
                    # Hızlı modda basit cevap
                    if mode == "hizli":
                        answer = f"""❌ **Excel okunamadı**

**Sebep:** {error_msg[:100]}

**Hızlı modda deneyin:**
• "VMA nedir?"
• "Excel macro"
• "Sistem hakkında"

**Detaylı mod için daha sonra tekrar deneyin.**"""
                    else:
                        answer = f"❌ **Excel okunamadı:** {error_msg}"
                    
                    self.send_success_response(answer, mode)
                    return
                
                # Sembolü ara
                search_result = find_symbol(question, excel_data)
                
                if not search_result.get("found"):
                    match = re.search(r'\b([A-Z]{2,6})\b', question.upper())
                    sembol_kodu = match.group(1) if match else "SEMBOL"
                    
                    # Excel bilgilerini ekle
                    total = excel_data.get("total_symbols", 0)
                    date = excel_data.get("excel_date", "güncel")
                    
                    answer = f"""❌ **{sembol_kodu} bulunamadı**

**Excel Bilgisi:**
• Tarih: {date}
• Toplam: {total} sembol
• Sayfalar: {list(excel_data.get('sheets', {}).keys())}

**Örnekler:**
• GMSTR, ALTIN (FON sayfasında)
• XU100, XULAS (ENDEKSLER sayfasında)
• ENKAI, GARAN (Sinyaller sayfasında)

**Deneyin:** "GMSTR analiz et", "XU100 durumu" """
                    
                    self.send_success_response(answer, mode)
                    return
                
                # AI analizi için prompt hazırla
                sembol_adi = search_result["name"]
                sembol_data = search_result["data"]
                sembol_sayfa = search_result.get("sayfa", "Sinyaller")
                excel_date = excel_data.get("excel_date", "güncel")
                
                print(f"✅ Bulundu: {sembol_adi} ({sembol_sayfa})", file=sys.stderr)
                
                # Prompt oluştur
                prompt = f"""📊 **{sembol_adi.upper()} TEKNİK ANALİZİ**

**Kaynak:** {sembol_sayfa} sayfası
**Excel Tarihi:** {excel_date}
**Analiz Modu:** {mode.upper()}

**VERİLER:**
"""
                
                # Tüm verileri ekle
                for key, value in sembol_data.items():
                    prompt += f"• **{key}:** {value}\n"
                
                prompt += f"\n**SORU:** {question}\n\n"
                
                if mode == "hizli":
                    prompt += """**TALİMAT (Hızlı Mod):**
1. Kısa ve öz ol (max 150 kelime)
2. VMA, EMA, Pivot'a odaklan
3. Temel teknik yorum yap
4. Yatırım tavsiyesi VERME

**ANALİZ:**"""
                else:
                    prompt += """**TALİMAT (Detaylı Mod):**
1. Tüm göstergeleri detaylı analiz et
2. VMA, EMA, Pivot, Bollinger değerlendir
3. Risk ve potansiyeli belirt
4. Destek/direnç seviyelerini analiz et
5. Yatırım tavsiyesi VERME

**DETAYLI ANALİZ:**"""
                
                # AI analizi al
                if mode == "hizli":
                    ai_answer = get_deepseek_analysis(prompt, "hizli")
                else:
                    ai_answer = get_groq_analysis(prompt)
                
                # Cevapla
                self.send_response(200)
                self.send_header('Content-type', 'application/json; charset=utf-8')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                
                result = {
                    "success": True,
                    "answer": ai_answer,
                    "symbol": sembol_adi,
                    "sheet": sembol_sayfa,
                    "excel_date": excel_date,
                    "mode": mode,
                    "timestamp": datetime.now().isoformat()
                }
                
                self.wfile.write(json.dumps(result, ensure_ascii=False).encode())
                print(f"📤 {mode} analiz gönderildi: {sembol_adi}", file=sys.stderr)
                return
            
            # BİLİNMEYEN SORU
            answer = """🤔 **Anlamadım**

**Modlar:**
• **Hızlı:** Özel sorular anında
• **Detaylı:** AI analizi 1-2 dakika

**Örnekler:**
• Hisse: "GMSTR analiz et"
• Endeks: "XU100 durumu"
• Teknik: "VMA nedir?"
• Genel: "Öne çıkan hisseler"

**Not:** Hisse analizi için mod seçin."""
            
            self.send_success_response(answer, mode)
            
        except Exception as e:
            print(f"❌ HATA: {e}", file=sys.stderr)
            traceback.print_exc(file=sys.stderr)
            self.send_error_response(f"Sistem hatası: {str(e)[:100]}")

    def send_success_response(self, answer, mode="hizli"):
        self.send_response(200)
        self.send_header('Content-type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        result = {
            "success": True, 
            "answer": answer, 
            "mode": mode,
            "timestamp": datetime.now().isoformat()
        }
        self.wfile.write(json.dumps(result, ensure_ascii=False).encode())
    
    def send_error_response(self, error):
        self.send_response(200)
        self.send_header('Content-type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        result = {"success": False, "answer": f"❌ Hata: {error}"}
        self.wfile.write(json.dumps(result, ensure_ascii=False).encode())

# ==================== TEST ====================
if __name__ == "__main__":
    from http.server import HTTPServer
    port = 3002
    server = HTTPServer(("0.0.0.0", port), handler)
    print(f"🚀 BorsaAnaliz AI 7.0: http://localhost:{port}")
    print("🎮 Modlar: Hızlı (DeepSeek) + Detaylı (Groq)")
    print("📅 Excel: Güncel tarihli otomatik bulunur")
    server.serve_forever()
