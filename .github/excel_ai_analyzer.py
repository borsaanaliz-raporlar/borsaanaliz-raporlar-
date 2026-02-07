#!/usr/bin/env python3
"""
M√úKEMMEL EXCEL ANALƒ∞ST - DEEPSEEK + GROQ Hƒ∞BRƒ∞T Sƒ∞STEM
"""
import os
import sys
import json
import re
from openpyxl import load_workbook
import requests
from datetime import datetime
from excel_finder import find_latest_excel

# AYARLAR
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY', '')

def get_excel_data_for_ai(excel_path):
    """AI i√ßin Excel verilerini hazƒ±rla - DETAYLI"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        all_data = {}
        
        target_sheets = ["Sinyaller", "ENDEKSLER", "FON_EMTIA_COIN_DOVIZ"]
        
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = []
                
                # T√úM ba≈ülƒ±klarƒ± al
                headers = []
                for col in range(1, 150):  # Daha fazla kolon
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(f"{cell_value}")
                    else:
                        break
                
                # T√úM hisseleri al (daha fazla)
                for row in ws.iter_rows(min_row=2, max_row=300, values_only=True):
                    if row and row[0]:
                        row_dict = {}
                        for i, cell_value in enumerate(row):
                            if i < len(headers):
                                row_dict[headers[i]] = cell_value
                        sheet_data.append(row_dict)
                
                all_data[sheet_name] = {
                    "headers": headers,
                    "data": sheet_data,
                    "row_count": len(sheet_data)
                }
                print(f"üìä {sheet_name}: {len(sheet_data)} hisse, {len(headers)} kolon")
        
        wb.close()
        
        return {
            "data": all_data,
            "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        }
        
    except Exception as e:
        return {"error": f"Excel okuma hatasƒ±: {str(e)}"}

def extract_hisse_adi(question):
    """Soru i√ßinden hisse adƒ±nƒ± √ßƒ±kar"""
    words = re.findall(r'\b[A-Z]{3,6}\b', question.upper())
    
    hisse_keywords = ["FROTO", "THYAO", "ASELS", "EREGL", "SASA", "KCHOL", 
                     "TOASO", "TUPRS", "AKBNK", "GARAN", "YKBNK", "XU100",
                     "GMSTR", "ALTIN", "XAUUSD", "XAGUSD", "XINSA", "XHOLD",
                     "XTEKS", "A1CAP", "ACSEL", "ADEL", "XU030"]
    
    for word in words:
        if word in hisse_keywords:
            return word
    
    if words:
        return words[0]
    
    return None

def get_hisse_analysis_data(hisse_info):
    """Hissenin analiz i√ßin gerekli T√úM verilerini √ßƒ±kar"""
    hisse_data = hisse_info["hisse"]
    headers = hisse_info["headers"]
    
    analysis_data = {
        "TEMEL": {},
        "Pƒ∞VOT_DESTEK_Dƒ∞REN√á": {},
        "HACƒ∞M": {},
        "EMA": {},
        "REGRESSION": {},
        "BOLLINGER": {},
        "Dƒ∞ƒûER": {}
    }
    
    for header in headers:
        if header in hisse_data:
            value = hisse_data[header]
            if value is None:
                continue
            
            header_upper = header.upper()
            
            if any(keyword in header_upper for keyword in ["Hƒ∞SSE", "SEMBOL", "CLOSE", "OPEN", "HIGH", "LOW"]):
                analysis_data["TEMEL"][header] = value
            elif any(keyword in header_upper for keyword in ["Pƒ∞VOT", "S1", "S2", "S3", "R1", "R2", "R3"]):
                analysis_data["Pƒ∞VOT_DESTEK_Dƒ∞REN√á"][header] = value
            elif any(keyword in header_upper for keyword in ["HACƒ∞M", "VOLUME"]):
                analysis_data["HACƒ∞M"][header] = value
            elif "EMA_" in header_upper:
                analysis_data["EMA"][header] = value
            elif any(keyword in header_upper for keyword in ["PEARSON", "KANAL", "UZAKLIK"]):
                analysis_data["REGRESSION"][header] = value
            elif header_upper.startswith("BB_"):
                analysis_data["BOLLINGER"][header] = value
            elif any(keyword in header_upper for keyword in ["VMA", "LSMA", "WT", "HMA", "SMI", "DURUM", "Sƒ∞NYAL"]):
                analysis_data["Dƒ∞ƒûER"][header] = value
    
    return analysis_data

def create_detailed_hisse_prompt(question, hisse_info, analysis_data):
    """DETAYLI hisse analizi i√ßin prompt"""
    
    hisse_name = hisse_info["hisse"].get(hisse_info["headers"][0], "HISSE")
    sheet_name = hisse_info["sheet"]
    
    prompt = f"""üéØ **SEN: BORSAANALIZ PROFESYONEL TEKNƒ∞K ANALƒ∞ST**

üìä **{hisse_name} DETAYLI TEKNƒ∞K ANALƒ∞Z RAPORU**

**Veri Kaynaƒüƒ±:** {sheet_name} sayfasƒ±
**Soru:** {question}

---

## üìà **1. TEMEL VERƒ∞LER:**
"""
    
    for key, value in analysis_data["TEMEL"].items():
        prompt += f"- **{key}:** {value}\n"
    
    prompt += f"""
## üìä **2. Pƒ∞VOT ve DESTEK/Dƒ∞REN√á ANALƒ∞Zƒ∞:**
"""
    
    for key, value in analysis_data["Pƒ∞VOT_DESTEK_Dƒ∞REN√á"].items():
        prompt += f"- **{key}:** {value}\n"
    
    close = analysis_data["TEMEL"].get("Close")
    pivot = analysis_data["Pƒ∞VOT_DESTEK_Dƒ∞REN√á"].get("Pivot")
    if close and pivot:
        try:
            close_f = float(str(close).replace(',', '.'))
            pivot_f = float(str(pivot).replace(',', '.'))
            if close_f > pivot_f:
                prompt += f"- **PIVOT ANALƒ∞Zƒ∞:** Fiyat pivotun √úST√úNDE (+%{((close_f-pivot_f)/pivot_f)*100:.2f})\n"
            else:
                prompt += f"- **PIVOT ANALƒ∞Zƒ∞:** Fiyat pivotun ALTINDA (-%{((pivot_f-close_f)/pivot_f)*100:.2f})\n"
        except:
            pass
    
    prompt += f"""
## üìä **3. HACƒ∞M ANALƒ∞Zƒ∞:**
"""
    
    for key, value in analysis_data["HACƒ∞M"].items():
        prompt += f"- **{key}:** {value}\n"
    
    prompt += f"""
## üìä **4. EMA (Exponential Moving Average) ANALƒ∞Zƒ∞:**
"""
    
    ema_items = sorted(analysis_data["EMA"].items())
    for key, value in ema_items:
        prompt += f"- **{key}:** {value}\n"
    
    close = analysis_data["TEMEL"].get("Close")
    if close and analysis_data["EMA"]:
        try:
            close_f = float(str(close).replace(',', '.'))
            for ema_key, ema_value in analysis_data["EMA"].items():
                try:
                    ema_f = float(str(ema_value).replace(',', '.'))
                    if "EMA_8" in ema_key and close_f > ema_f:
                        prompt += f"- **EMA_8 YORUM:** Fiyat EMA_8'in √úST√úNDE (Kƒ±sa vadeli trend POZƒ∞Tƒ∞F)\n"
                        break
                except:
                    pass
        except:
            pass
    
    prompt += f"""
## üìä **5. REGRESSION KANAL ANALƒ∞Zƒ∞:**
"""
    
    for key, value in analysis_data["REGRESSION"].items():
        prompt += f"- **{key}:** {value}\n"
    
    pearson55 = analysis_data["REGRESSION"].get("Pearson55")
    if pearson55:
        try:
            p55 = float(str(pearson55).replace(',', '.'))
            if p55 > 0.3:
                prompt += f"- **55 G√úN REGRESSION:** Pearson={p55:.3f} > 0.3 = Y√úKSELƒ∞≈û TRENDƒ∞\n"
            elif p55 < -0.3:
                prompt += f"- **55 G√úN REGRESSION:** Pearson={p55:.3f} < -0.3 = D√ú≈û√ú≈û TRENDƒ∞\n"
            else:
                prompt += f"- **55 G√úN REGRESSION:** Pearson={p55:.3f} = N√ñTR/RANGE\n"
        except:
            pass
    
    prompt += f"""
## üìä **6. BOLLINGER BANDS ANALƒ∞Zƒ∞:**
"""
    
    for key, value in analysis_data["BOLLINGER"].items():
        prompt += f"- **{key}:** {value}\n"
    
    bb_upper = analysis_data["BOLLINGER"].get("BB_UPPER")
    bb_middle = analysis_data["BOLLINGER"].get("BB_MIDDLE")
    bb_lower = analysis_data["BOLLINGER"].get("BB_LOWER")
    close = analysis_data["TEMEL"].get("Close")
    
    if all([bb_upper, bb_middle, bb_lower, close]):
        try:
            close_f = float(str(close).replace(',', '.'))
            upper_f = float(str(bb_upper).replace(',', '.'))
            lower_f = float(str(bb_lower).replace(',', '.'))
            
            if close_f > upper_f:
                prompt += f"- **BOLLINGER YORUM:** Fiyat √ºst bandƒ±n √úST√úNDE = A≈ûIRI ALIM\n"
            elif close_f < lower_f:
                prompt += f"- **BOLLINGER YORUM:** Fiyat alt bandƒ±n ALTINDA = A≈ûIRI SATIM\n"
            else:
                prompt += f"- **BOLLINGER YORUM:** Fiyat bantlar ƒ∞√áƒ∞NDE = NORMAL\n"
        except:
            pass
    
    prompt += f"""
## üìä **7. Dƒ∞ƒûER TEKNƒ∞K G√ñSTERGELER:**
"""
    
    for key, value in analysis_data["Dƒ∞ƒûER"].items():
        prompt += f"- **{key}:** {value}\n"
    
    vma_value = None
    for key, value in analysis_data["Dƒ∞ƒûER"].items():
        if "VMA" in key.upper():
            vma_value = str(value)
            break
    
    if vma_value:
        if "POZƒ∞Tƒ∞F" in vma_value.upper():
            prompt += f"- **VMA YORUM:** {vma_value} = Hacim trendi Y√úKSELƒ∞≈û (%94 doƒüruluk)\n"
        elif "NEGATƒ∞F" in vma_value.upper():
            prompt += f"- **VMA YORUM:** {vma_value} = Hacim trendi D√ú≈û√ú≈û (%94 doƒüruluk)\n"
    
    prompt += f"""

---

## üìã **TEKNƒ∞K ANALƒ∞Z TALƒ∞MATLARI:**

**MUTLAKA YAP:**
1. Yukarƒ±daki T√úM verilere g√∂re detaylƒ± analiz yap
2. Her b√∂l√ºm√º tek tek deƒüerlendir
3. Sayƒ±sal verileri YORUMLA
4. Trendleri belirle
5. Risk seviyesini deƒüerlendir

**YAPMA:**
1. ASLA "Volkswagen" deme (VMA = Volume Moving Algorithm)
2. RSI/MACD'den bahsetme (yok!)
3. Yatƒ±rƒ±m tavsiyesi verme

**ANALƒ∞Z BA≈ûLIKLARI:**
1. Genel Teknik Durum √ñzeti
2. Pivot ve Destek/Diren√ß Analizi
3. Hacim Analizi
4. EMA Trend Analizi
5. Regression Kanal Analizi
6. Bollinger Bands Analizi
7. VMA ve Diƒüer G√∂stergeler
8. Risk Deƒüerlendirmesi

---

**≈ûƒ∞MDƒ∞ {hisse_name} ƒ∞√áƒ∞N DETAYLI TEKNƒ∞K ANALƒ∞Z YAP:**
"""
    
    return prompt

def create_general_prompt(question, excel_data):
    """Genel analiz i√ßin prompt"""
    
    data = excel_data["data"]
    timestamp = excel_data["timestamp"]
    
    prompt = f"""üéØ **SEN: BORSAANALIZ PROFESYONEL TEKNƒ∞K ANALƒ∞ST**

üìä **Pƒ∞YASA ANALƒ∞Z RAPORU**
**Tarih:** {timestamp}
**Soru:** {question}

---

## üìà **ELƒ∞MDEKƒ∞ VERƒ∞LER:**

"""

    for sheet_name, sheet_info in data.items():
        headers = sheet_info["headers"]
        hisse_sayisi = len(sheet_info["data"])
        
        prompt += f"""
### {sheet_name.upper()} SAYFASI:
‚Ä¢ **Hisse Sayƒ±sƒ±:** {hisse_sayisi}
‚Ä¢ **Kolon Sayƒ±sƒ±:** {len(headers)}
‚Ä¢ **√ñnemli G√∂stergeler:**"""
        
        important_indicators = []
        for header in headers:
            if any(keyword in header.upper() for keyword in 
                   ["VMA", "LSMA", "WT", "EMA", "PEARSON", "KANAL", 
                    "PIVOT", "HACƒ∞M", "BB_", "HMA", "SMI"]):
                important_indicators.append(header)
        
        prompt += f" {', '.join(important_indicators[:10])}"
        if len(important_indicators) > 10:
            prompt += f" ..."
        
        prompt += f"\n‚Ä¢ **ƒ∞lk 5 Hisse:** "
        hisse_list = []
        for hisse in sheet_info["data"][:5]:
            hisse_name = hisse.get(headers[0], "")
            if hisse_name:
                hisse_list.append(hisse_name)
        prompt += f"{', '.join(hisse_list)}"
    
    prompt += f"""

---

## üìã **ANALƒ∞Z TALƒ∞MATLARI:**

**SADECE yukarƒ±daki verileri kullanarak:**
1. {question} sorusunu cevapla
2. Hisse isimlerini GER√áEK yaz
3. Teknik g√∂stergeleri doƒüru kullan
4. Regression sorulursa: Pearson55, Pearson144, Pearson233 kontrol et
5. VMA = Volume Moving Algorithm (%94 doƒüruluk)

**YAPMA:**
1. RSI/MACD deme (yok!)
2. Yatƒ±rƒ±m tavsiyesi verme

---

**CEVAP FORMATI:**
1. üìä Analiz √ñzeti
2. üìà Teknik Bulgular
3. üîç Detaylƒ± Analiz
4. ‚ö†Ô∏è Risk Uyarƒ±sƒ±

---

**≈ûƒ∞MDƒ∞ ANALƒ∞Z YAP:**
"""
    
    return prompt

def call_deepseek_ai(prompt, question):
    """DeepSeek AI √ßaƒüƒ±r"""
    if not DEEPSEEK_API_KEY:
        return "DEEPSEEK_API_KEY eksik"
    
    data = {
        "model": "deepseek-chat",
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user", "content": question}
        ],
        "max_tokens": 2000,
        "temperature": 0.1,
        "stream": False
    }
    
    try:
        response = requests.post(
            "https://api.deepseek.com/chat/completions",
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}"},
            json=data,
            timeout=60
        )
        
        if response.status_code == 200:
            answer = response.json()['choices'][0]['message']['content']
            
            # Kontroller
            answer_lower = answer.lower()
            
            if "volkswagen" in answer_lower:
                answer = answer.replace("Volkswagen", "Volume Moving Algorithm")
            
            if "rsi" in answer_lower or "macd" in answer_lower:
                answer += "\n\n‚ö†Ô∏è **NOT:** Excel'de RSI ve MACD g√∂stergeleri bulunmamaktadƒ±r."
            
            if "yatƒ±rƒ±m tavsiyesi deƒüildir" not in answer_lower:
                answer += "\n\n‚ö†Ô∏è **√ñNEMLƒ∞ UYARI:** Bu analiz bilgi ama√ßlƒ±dƒ±r, yatƒ±rƒ±m tavsiyesi deƒüildir. Yatƒ±rƒ±m kararlarƒ±nƒ±zƒ± kendi ara≈ütƒ±rmanƒ±zla alƒ±nƒ±z."
            
            return answer
        else:
            return f"‚ùå DeepSeek API hatasƒ±: {response.status_code}"
            
    except Exception as e:
        return f"‚ùå DeepSeek baƒülantƒ± hatasƒ±: {str(e)}"

def call_groq_ai(prompt, question):
    """GROQ AI √ßaƒüƒ±r"""
    if not GROQ_API_KEY:
        return "GROQ_API_KEY eksik"
    
    data = {
        "model": "llama-3.3-70b-versatile",
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user", "content": "L√ºtfen detaylƒ± teknik analiz yap."}
        ],
        "max_tokens": 2000,
        "temperature": 0.1,
        "top_p": 0.9,
        "stream": False
    }
    
    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
            json=data,
            timeout=90
        )
        
        if response.status_code == 200:
            answer = response.json()['choices'][0]['message']['content']
            
            # Kontroller
            answer_lower = answer.lower()
            
            if "volkswagen" in answer_lower:
                answer = answer.replace("Volkswagen", "Volume Moving Algorithm")
            
            if "rsi" in answer_lower or "macd" in answer_lower:
                answer += "\n\n‚ö†Ô∏è **NOT:** Excel'de RSI ve MACD g√∂stergeleri bulunmamaktadƒ±r."
            
            if "yatƒ±rƒ±m tavsiyesi deƒüildir" not in answer_lower:
                answer += "\n\n‚ö†Ô∏è **√ñNEMLƒ∞ UYARI:** Bu analiz bilgi ama√ßlƒ±dƒ±r, yatƒ±rƒ±m tavsiyesi deƒüildir. Yatƒ±rƒ±m kararlarƒ±nƒ±zƒ± kendi ara≈ütƒ±rmanƒ±zla alƒ±nƒ±z."
            
            return answer
        else:
            return f"‚ùå GROQ API hatasƒ±: {response.status_code}"
            
    except Exception as e:
        return f"‚ùå GROQ baƒülantƒ± hatasƒ±: {str(e)}"

def call_ai_analyst(prompt, question, use_deepseek=True):
    """Hƒ∞BRƒ∞T AI √ßaƒüƒ±rƒ±cƒ± - √ñnce DeepSeek, olmazsa GROQ"""
    
    if use_deepseek and DEEPSEEK_API_KEY:
        print("üöÄ DeepSeek AI kullanƒ±lƒ±yor...")
        answer = call_deepseek_ai(prompt, question)
        
        # Eƒüer DeepSeek ba≈üarƒ±lƒ±ysa d√∂n
        if "hatasƒ±" not in answer and len(answer) > 100:
            return answer
        else:
            print(f"‚ö†Ô∏è DeepSeek √ßalƒ±≈ümadƒ±, GROQ'a ge√ßiliyor: {answer[:100]}")
    
    # DeepSeek yoksa veya √ßalƒ±≈ümazsa GROQ
    print("‚ö° GROQ AI kullanƒ±lƒ±yor...")
    return call_groq_ai(prompt, question)

def main():
    """Ana fonksiyon"""
    if len(sys.argv) < 2:
        print("‚ùå Kullanƒ±m: python excel_ai_analyzer.py 'SORUNUZ'")
        return
    
    question = sys.argv[1]
    print(f"‚ùì Soru: {question}")
    
    print("üîç Excel dosyasƒ± aranƒ±yor...")
    excel_info = find_latest_excel()
    
    if not excel_info:
        return "‚ö†Ô∏è Excel dosyasƒ± bulunamadƒ±"
    
    print(f"üìñ Excel: {excel_info['name']}")
    
    # Excel verilerini al
    excel_data = get_excel_data_for_ai(excel_info['path'])
    
    if "error" in excel_data:
        answer = f"‚ùå {excel_data['error']}"
    else:
        # Hisse analizi mi?
        hisse_adi = extract_hisse_adi(question)
        
        if hisse_adi:
            print(f"üéØ Hisse analizi: {hisse_adi}")
            
            # Hisseyi bul
            hisse_info = None
            for sheet_name, sheet_info in excel_data["data"].items():
                headers = sheet_info["headers"]
                for hisse in sheet_info["data"]:
                    hisse_name = hisse.get(headers[0], "")
                    if hisse_name and hisse_adi in str(hisse_name).upper():
                        hisse_info = {
                            "hisse": hisse,
                            "headers": headers,
                            "sheet": sheet_name
                        }
                        break
                if hisse_info:
                    break
            
            if hisse_info:
                # DETAYLI analiz verilerini hazƒ±rla
                analysis_data = get_hisse_analysis_data(hisse_info)
                
                # DETAYLI prompt olu≈ütur
                prompt = create_detailed_hisse_prompt(question, hisse_info, analysis_data)
                answer = call_ai_analyst(prompt, question)
            else:
                answer = f"‚ùå {hisse_adi} hissesi Excel'de bulunamadƒ±"
        else:
            # Genel analiz
            prompt = create_general_prompt(question, excel_data)
            answer = call_ai_analyst(prompt, question)
    
    # Sonucu kaydet
    with open('ai_response.txt', 'w', encoding='utf-8') as f:
        f.write(answer)
    
    print("\n" + "="*60)
    print("‚úÖ ANALƒ∞Z TAMAMLANDI!")
    print("="*60)
    print(f"\nüìä AI YANITI:\n")
    print(answer[:800] + "..." if len(answer) > 800 else answer)
    print("\n" + "="*60)
    
    return answer

if __name__ == "__main__":
    main()
