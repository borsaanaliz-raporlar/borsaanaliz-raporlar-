#!/usr/bin/env python3
"""
BORSAANALİZ V11 UZMAN TEKNİK ANALİST
⚡ HIZLI (Groq) + 📋 DETAYLI (Groq) BUTONLU SİSTEM
AI'IN VERİ UYDURMASI ENGELLENDİ!
"""
import os
import sys
import json
import re
from openpyxl import load_workbook
import requests
from datetime import datetime
from excel_finder import find_latest_excel

# API AYARLARI
DEEPSEEK_API_KEY = os.environ.get('DEEPSEEK_API_KEY', '')
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')

# ============= MODE AYARI =============
def is_detailed_mode(question):
    """Detaylı analiz isteği kontrolü"""
    detailed_keywords = ["detaylı", "detayli", "kapsamlı", "kapsamli", "profesyonel", "uzun"]
    return any(keyword in question.lower() for keyword in detailed_keywords)

def get_excel_data_for_ai(excel_path):
    """Excel'deki TÜM verileri al"""
    try:
        wb = load_workbook(excel_path, data_only=True, read_only=True)
        all_data = {}
        
        target_sheets = ["Sinyaller", "ENDEKSLER", "FON_EMTIA_COIN_DOVIZ"]
        
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = []
                
                headers = []
                col = 1
                while True:
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value).strip())
                        col += 1
                    else:
                        break
                
                for row in ws.iter_rows(min_row=2, max_row=500, values_only=True):
                    if row and row[0] and row[0] not in ["Toplam", "Genel", "Ortalama", "Sektör", "Hisse", "Sembol"]:
                        row_dict = {}
                        for i, val in enumerate(row):
                            if i < len(headers) and val is not None:
                                row_dict[headers[i]] = val
                        if row_dict:
                            sheet_data.append(row_dict)
                
                all_data[sheet_name] = {
                    "headers": headers,
                    "data": sheet_data,
                    "count": len(sheet_data)
                }
                print(f"✅ {sheet_name}: {len(sheet_data)} hisse, {len(headers)} kolon")
        
        wb.close()
        
        return {
            "data": all_data,
            "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            "file": os.path.basename(excel_path)
        }
        
    except Exception as e:
        return {"error": f"Excel okuma hatası: {str(e)}"}

def extract_hisse_adi(question):
    """HİSSE ADI BULUCU - LİSTE YOK, SAF REGEX!"""
    words = re.findall(r'\b[A-Z0-9]{3,8}\b', question.upper())
    return words[0] if words else None

def get_hisse_raw_data(hisse_info, headers):
    """Hisse'nin ham verilerini formatlı şekilde döndür"""
    ham_veriler = ""
    kritik_kolonlar = [
        "Hisse", "Close", "Pivot", "WT Sinyal", "WT1", "WT2",
        "VMA trend algo", "LSMA KAMA", "HMA_89",
        "S3", "S2", "S1", "R1", "R2", "R3",
        "EMA_8", "EMA_13", "EMA_21", "EMA_34", "EMA_55", "EMA_89", "EMA_144", "EMA_233",
        "Pearson55", "Pearson89", "Pearson144", "Pearson233",
        "BB_UPPER", "BB_MIDDLE", "BB_LOWER",
        "Hacim", "Hacim_Değişim_%", "Hacim_Senaryo",
        "SMI", "SMI_EMA", "AI_YORUM"
    ]
    
    for kolon in kritik_kolonlar:
        if kolon in hisse_info and hisse_info[kolon] is not None:
            deger = hisse_info[kolon]
            # Sayısal değerleri formatla
            if isinstance(deger, (int, float)):
                if kolon in ["Close", "Pivot", "S1", "S2", "S3", "R1", "R2", "R3"]:
                    ham_veriler += f"• **{kolon}:** {deger:.2f}\n"
                else:
                    ham_veriler += f"• **{kolon}:** {deger}\n"
            else:
                ham_veriler += f"• **{kolon}:** {deger}\n"
    
    return ham_veriler

def create_quick_prompt(question, excel_data, hisse_adi=None):
    """⚡ HIZLI ANALİZ - Kısa prompt"""
    
    timestamp = excel_data["timestamp"]
    data = excel_data["data"]
    
    system_intro = f"""🎯 **BORSAANALİZ V11 HIZLI ANALİZ**
📅 {timestamp}

**📊 KRİTİK GÖSTERGELER:**
• VMA: %94 doğruluk, parantez içi GÜN SAYISI
• POZİTİF(57) = 57 gündür yükselen trend, NEGATİF(7) = 7 gündür düşen trend
• LSMA: Trend göstergesi, parantez içi GÜN SAYISI
• Pearson: >0.3 yükseliş, <-0.3 düşüş

**🚫 YOK:** RSI, MACD, Stokastik
"""
    
    if hisse_adi:
        hisse_info = None
        sheet_name = None
        found = False
        
        for sname, sinfo in data.items():
            for hisse in sinfo["data"]:
                hisse_name = hisse.get(sinfo["headers"][0], "")
                if hisse_name and hisse_adi.upper() in str(hisse_name).upper():
                    hisse_info = hisse
                    sheet_name = sname
                    found = True
                    break
            if found:
                break
        
        if hisse_info:
            ham_veri = get_hisse_raw_data(hisse_info, sinfo["headers"])
            
            # Veri var mı kontrol et
            if not ham_veri.strip():
                return f"⚠️ **{hisse_adi}** için Excel'de veri bulunamadı. Lütfen sembolü kontrol edin."
            
            prompt = system_intro + f"""

📊 **{hisse_adi} HAM VERİLER (SADECE BUNLAR GERÇEK):**
{ham_veri}

**⚠️ ÇOK ÖNEMLİ UYARI:**
Yukarıdaki HAM VERİLER dışında hiçbir rakam KULLANMA!
Eğer bir bilgi yukarıda yoksa, SAKIN kendin rakam uydurma!

**ŞU SORULARA CEVAP VER:**
1. Kısa vadeli görünüm (EMA8/21, WT)
2. Destek/direnç seviyeleri (S1/R1)
3. VMA trendi kaç gün? Ne anlama gelir?
4. Pearson regresyon analizi
5. Hacim senaryosu yorumu

⚠️ Yatırım tavsiyesi değildir.
"""
            return prompt
        else:
            return f"⚠️ **{hisse_adi}** Excel dosyasında bulunamadı. Lütfen sembolü kontrol edin."
    
    # Genel analiz
    return system_intro + f"""

📊 **PİYASA GENEL GÖRÜNÜM**

**Soru:** {question}

Hızlı piyasa analizi yap:
- Endekslerin durumu
- Öne çıkan hisseler
- Genel trend yönü
"""

def create_detailed_prompt(question, excel_data, hisse_adi=None):
    """📋 DETAYLI ANALİZ - Uzun prompt"""
    
    timestamp = excel_data["timestamp"]
    data = excel_data["data"]
    
    system_intro = f"""🎯 **BORSAANALİZ V11 PROFESYONEL ANALİST**
📅 {timestamp}

═══════════════════════════════════════════
**📌 GÖSTERGE YORUM KILAVUZU:**

1️⃣ **WT (WaveTrend):** 
   • >60 = Aşırı alım
   • <-60 = Aşırı satım
   • POZİTİF/NEGATİF = Trend yönü

2️⃣ **VMA (hacim ağırlıklı trend algoritmasıdır):**
   • VMA %94 doğrulukla sinyal üreten özel bir algoritmadır
   • Bu, basit bir hareketli ortalama DEĞİLDİR!
   • POZİTİF(57) = 57 gündür yükselen trend devam ediyor
   • NEGATİF(7) = 7 gündür düşen trend devam ediyor
   • ASLA "Volume Moving Average" olarak yorumlama!
   • ASLA fiyatla karşılaştırma, sadece HACİM AĞIRLIKLI TREND olarak yorumla!

3️⃣ **LSMA KAMA:**
   • POZİTİF(48) = 48 gündür yükseliş trendi devam ediyor 
   • NEGATİF(5) = 5 gündür düşüş trendi devam ediyor

4️⃣ **PEARSON KATSAYISI:**
   • 0.70-1.00 = ÇOK GÜÇLÜ trend
   • 0.30-0.70 = GÜÇLÜ trend
   • 0.10-0.30 = ZAYIF trend
   • -0.10-0.10 = YATAY/BELİRSİZ

5️⃣ **EMA HİYERARŞİSİ:**
   • 8>13>21 = YÜKSELİŞ
   • 8<13<21 = DÜŞÜŞ
   • Karmaşık = YATAY

6️⃣ **HACİM SENARYOLARI:**
   • POZITIF_YUKSELME = Hacim artışıyla yükseliş (GÜVENİLİR)
   • NEGATIF_DUSUS = Hacim düşüşüyle düşüş (GÜVENİLİR)
   • POZITIF_DUSUS = Hacim artışıyla düşüş (SATIŞ BASKISI)
   • NEGATIF_YUKSELME = Hacim düşüşüyle yükseliş (ZAYIF)

7️⃣ **BOLLINGER BANTLARI:**
   • Fiyat > Üst bant = AŞIRI ALIM
   • Fiyat < Alt bant = AŞIRI SATIM
   • Bant içinde = NORMAL

**🚫 KESİNLİKLE YOK:** RSI, MACD, Stokastik - SAKIN KULLANMA!
**📌 PARANTEZ İÇİNDEKİ RAKAMLAR:** Trendin kaç gündür devam ettiğini gösterir!
═══════════════════════════════════════════
"""
    
    if hisse_adi:
        hisse_info = None
        sheet_name = None
        found = False
        
        for sname, sinfo in data.items():
            for hisse in sinfo["data"]:
                hisse_name = hisse.get(sinfo["headers"][0], "")
                if hisse_name and hisse_adi.upper() in str(hisse_name).upper():
                    hisse_info = hisse
                    sheet_name = sname
                    found = True
                    break
            if found:
                break
        
        if hisse_info:
            ham_veri = get_hisse_raw_data(hisse_info, sinfo["headers"])
            
            # Veri var mı kontrol et
            if not ham_veri.strip():
                return f"⚠️ **{hisse_adi}** için Excel'de veri bulunamadı. Lütfen sembolü kontrol edin."
            
            prompt = system_intro + f"""

═══════════════════════════════════════════
📋 **DETAYLI ANALİZ: {hisse_adi}**
📌 **Kaynak:** {sheet_name}
═══════════════════════════════════════════

**📊 HAM VERİLER (SADECE BUNLAR GERÇEK):**
{ham_veri}

**⚠️ ÇOK ÖNEMLİ UYARI:**
Yukarıdaki HAM VERİLER dışında hiçbir rakam KULLANMA!
Eğer bir bilgi yukarıda yoksa, SAKIN kendin rakam uydurma!

**🔍 ŞU BAŞLIKLARDA DETAYLI ANALİZ YAP:**

1️⃣ **KISA VADELİ GÖRÜNÜM (1-5 GÜN)**
   • WT sinyali ve seviyesi
   • EMA8/EMA21 ilişkisi
   • VMA gün sayısı yorumu
   • İlk hedef direnç (R1)

2️⃣ **ORTA VADELİ GÖRÜNÜM (1-4 HAFTA)**
   • LSMA trend süresi (kaç gün?)
   • Pearson55/89 değerleri ve gücü
   • EMA hiyerarşisi analizi
   • Ana trend yönü

3️⃣ **KRİTİK SEVİYELER**
   • S1-R1 günlük hareket bandı
   • S3 (stop-loss bölgesi)
   • R3 (hedef bölgesi)
   • Pivot'a göre konum

4️⃣ **HACİM ANALİZİ**
   • VMA trendi ve süresi
   • Hacim senaryosu yorumu
   • Hacim değişim yüzdesi
   • Güvenilirlik değerlendirmesi

5️⃣ **REGRESYON ANALİZİ**
   • Pearson55 trend gücü
   • Kanal üst/alt seviyeleri
   • Fiyatın kanaldaki konumu

6️⃣ **RİSK DEĞERLENDİRMESİ**
   • Düşük/Orta/Yüksek
   • Nedenleriyle açıkla
   • Volatilite durumu

7️⃣ **YATIRIMCI NOTU**
   • İzlenecek seviyeler
   • Olası senaryolar
   • Strateji önerisi

⚠️ **YASAL UYARI:** Yatırım tavsiyesi değildir.
"""
            return prompt
        else:
            return f"⚠️ **{hisse_adi}** Excel dosyasında bulunamadı. Lütfen sembolü kontrol edin."
    
    # Genel detaylı analiz
    return system_intro + f"""

═══════════════════════════════════════════
📋 **PİYASA DETAYLI ANALİZ**
═══════════════════════════════════════════

**Soru:** {question}

**🔍 DETAYLI PİYASA ANALİZİ:**

1️⃣ **ENDERSLERİN TEKNİK DURUMU**
   • XU100, XU030, XBANK analizi
   • WT sinyalleri
   • VMA trendleri

2️⃣ **ÖNE ÇIKAN HİSSELER**
   • En uzun VMA POZİTİF olanlar
   • Pearson55 > 0.85 olanlar
   • Hacim senaryosu güçlü olanlar

3️⃣ **SEKTÖREL DEĞERLENDİRME**
   • En güçlü sektör endeksleri
   • En zayıf sektör endeksleri
   • Sektör rotasyonu var mı?

4️⃣ **RİSK İŞTAHI**
   • POZITIF_YUKSELME oranı
   • NEGATIF_DUSUS oranı
   • Genel piyasa hissiyatı

⚠️ **YASAL UYARI:** Yatırım tavsiyesi değildir.
"""

def call_groq(prompt, question):
    """Groq AI çağrısı - ANA ANALİZ MOTORU - VERİ UYDURMA ENGELLİ"""
    if not GROQ_API_KEY:
        return None
    
    try:
        print("⚡ Groq AI analiz yapıyor...")
        
        # Prompt'ta gerçek veri var mı kontrol et
        if "HAM VERİLER" not in prompt and "Close" not in prompt:
            print("⚠️ UYARI: Prompt'ta ham veri yok!")
            return "⚠️ Excel'de bu sorgu için yeterli veri bulunamadı."
        
        # Groq için özel sistem mesajı - ÇOK SIKI KURALLAR
        system_message = """Sen BORSAANALİZ V11 uzmanısın.

🚫 **KESİNLİKLE YASAK OLANLAR:**
1. SAKIN kendin rakam uydurma!
2. SAKIN "POZİTİF(57)" gibi örnek rakamlar kullanma!
3. SAKIN "LSMA: 20 gün" gibi uydurma değerler yazma!
4. SAKIN "Pearson: 0.5" gibi uydurma katsayılar yazma!

✅ **SADECE ŞUNLARI YAP:**
1. Sana verilen "HAM VERİLER" başlığı altındaki bilgileri KULLAN
2. Eğer bir bilgi HAM VERİLER'de yoksa "Bu bilgi Excel'de bulunmamaktadır" DE
3. Sadece gerçek verilerle yorum yap

📋 **ÖRNEK DOĞRU CEVAP:**
"AKBNK için Excel verilerinde Close: 15.34, VMA trend algo: NEGATİF(7) olarak görünüyor. Destek seviyeleri S1:14.90, S2:14.50..."

📋 **ÖRNEK YANLIŞ CEVAP (SAKIN YAPMA):**
"VMA: POZİTİF(57), LSMA: 20 gün, Pearson: 0.5" (çünkü bu rakamlar uydurma!)

Şimdi aşağıdaki verilere göre cevap ver:"""
        
        # Kullanıcı mesajını hazırla
        user_message = f"""📊 **EXCEL'DEN ALINAN GERÇEK VERİLER (SADECE BUNLAR VAR):**
{prompt}

❓ **SORU:** {question}

⚠️ **SON UYARI:** Yukarıdaki veriler dışında hiçbir rakam kullanma!
Eğer bir bilgi yoksa "Bu bilgi Excel'de bulunmamaktadır" de."""
        
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "messages": [
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": user_message}
                ],
                "temperature": 0.1,  # Çok düşük, neredeyse hiç yaratıcılık yok
                "max_tokens": 2000
            },
            timeout=60
        )
        
        if response.status_code == 200:
            answer = response.json()['choices'][0]['message']['content']
            
            # Uydurma rakam kontrolü
            uydurma_kontrol = [
                "POZİTİF(57)", "NEGATİF(57)", "POZİTİF(48)", "NEGATİF(48)",
                "LSMA: 20", "LSMA: 15", "LSMA: 10", "LSMA: 5",
                "Pearson: 0.5", "Pearson: 0.6", "Pearson: 0.7", "Pearson: 0.8",
                "Pearson: -0.5", "Pearson: -0.6"
            ]
            
            for uydurma in uydurma_kontrol:
                if uydurma in answer and uydurma not in prompt:
                    print(f"⚠️ UYARI: AI '{uydurma}' uydurdu! Cevap düzeltiliyor...")
                    answer = f"⚠️ **UYARI: AI veri uydurmaya çalıştı!**\n\nLütfen Excel dosyasını kontrol edin.\n\n{answer}"
                    break
            
            # Yazım hatalarını düzelt
            answer = answer.replace("Volume Moving Average", "HACİM AĞIRLIKLI TREND ALGORİTMASI")
            answer = answer.replace("Volumetric Moving Average", "HACİM AĞIRLIKLI TREND ALGORİTMASI")
            answer = answer.replace("Garanti Bankası", "QNB Finansbank GÜMÜŞ FONU")
            answer = answer.replace("Hacım", "Hacim")
            
            print(f"✅ Groq başarılı!")
            return answer
        else:
            print(f"⚠️ Groq hata {response.status_code}")
            return None
            
    except Exception as e:
        print(f"⚠️ Groq bağlantı hatası: {str(e)}")
        return None

def main():
    """Ana fonksiyon - SADECE GROQ KULLANIR, veri uydurma engelli"""
    if len(sys.argv) < 2:
        print("❌ Hata: Soru girmediniz!")
        return
    
    question = sys.argv[1]
    print(f"❓ SORU: {question}")
    
    # MOD BELİRLEME
    detailed_mode = is_detailed_mode(question)
    print(f"📋 MOD: {'DETAYLI' if detailed_mode else 'HIZLI'}")
    
    # Excel bul
    print("🔍 Excel dosyası aranıyor...")
    excel_info = find_latest_excel()
    
    if not excel_info:
        print("❌ Excel dosyası bulunamadı!")
        answer = "⚠️ Excel dosyası bulunamadı. Lütfen raporlar/ klasörünü kontrol edin."
        
        with open('ai_response.txt', 'w', encoding='utf-8') as f:
            f.write(answer)
        return
    
    print(f"📁 Excel: {excel_info['name']}")
    
    # Excel verilerini oku
    excel_data = get_excel_data_for_ai(excel_info['path'])
    
    if "error" in excel_data:
        answer = f"❌ {excel_data['error']}"
    else:
        hisse_adi = extract_hisse_adi(question)
        
        if hisse_adi:
            print(f"🎯 Hisse: {hisse_adi}")
        else:
            print("📊 Genel piyasa analizi")
        
        # MOD'A GÖRE PROMPT OLUŞTUR
        if detailed_mode:
            prompt = create_detailed_prompt(question, excel_data, hisse_adi)
        else:
            prompt = create_quick_prompt(question, excel_data, hisse_adi)
        
        # Eğer prompt zaten bir hata mesajıysa (veri bulunamadı gibi), onu direkt kullan
        if prompt.startswith("⚠️"):
            answer = prompt
        else:
            # Groq ile analiz yap
            print("⚡ Groq ile analiz yapılıyor...")
            answer = call_groq(prompt, question)
        
        # HİÇBİRİ ÇALIŞMAZSA
        if not answer:
            answer = f"""⚠️ **AI SERVİSİNE ULAŞILAMADI**

📁 Excel: {excel_info['name']}
📅 Tarih: {excel_data['timestamp']}
📋 Mod: {'DETAYLI' if detailed_mode else 'HIZLI'}

Lütfen API anahtarını kontrol edin."""
    
    # Yanıtı kaydet
    with open('ai_response.txt', 'w', encoding='utf-8') as f:
        f.write(answer)
    
    print(f"\n✅ ANALİZ TAMAMLANDI!")
    print(f"📝 Yanıt kaydedildi: ai_response.txt")
    print(f"📏 Yanıt uzunluğu: {len(answer)} karakter")

if __name__ == "__main__":
    main()
